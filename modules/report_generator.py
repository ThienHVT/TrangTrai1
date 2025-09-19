from openpyxl import Workbook
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

class ReportGenerator:
    @staticmethod
    def generate_excel_report(data, report_type, filename=None):
        """
        Tạo báo cáo Excel từ dữ liệu
        
        Args:
            data: Dữ liệu cần xuất (list of dict)
            report_type: Loại báo cáo ('crops', 'animals', 'activities')
            filename: Tên file output (nếu None sẽ tự động tạo)
        
        Returns:
            Đường dẫn file đã tạo
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"bao_cao_{report_type}_{timestamp}.xlsx"
        
        # Tạo workbook mới
        wb = Workbook()
        ws = wb.active
        
        # Đặt tiêu đề worksheet
        report_titles = {
            'crops': 'BÁO CÁO CÂY TRỒNG',
            'animals': 'BÁO CÁO VẬT NUÔI',
            'activities': 'BÁO CÁO HOẠT ĐỘNG'
        }
        ws.title = report_titles.get(report_type, "BÁO CÁO")
        
        # Định dạng cơ bản
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Xác định headers và columns dựa trên loại báo cáo
        if report_type == 'crops':
            headers = ["ID", "Tên Cây", "Loại Cây", "Ngày Trồng", "Diện Tích (ha)", "Trạng Thái", "Ghi Chú"]
            columns = ["id", "name", "type", "planting_date", "area", "status", "notes"]
            col_widths = [8, 25, 20, 15, 15, 15, 40]
        elif report_type == 'animals':
            headers = ["ID", "Tên Vật Nuôi", "Loại", "Ngày Nhập", "Số Lượng", "Trạng Thái", "Ghi Chú"]
            columns = ["id", "name", "type", "entry_date", "quantity", "status", "notes"]
            col_widths = [8, 25, 20, 15, 15, 15, 40]
        elif report_type == 'activities':
            headers = ["ID", "Tên Hoạt Động", "Loại", "Ngày", "Người Phụ Trách", "Trạng Thái", "Mô Tả"]
            columns = ["id", "name", "type", "date", "responsible", "status", "description"]
            col_widths = [8, 25, 20, 15, 20, 15, 40]
        else:
            raise ValueError("Loại báo cáo không hợp lệ")
        
        # Thêm tiêu đề báo cáo
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = report_titles.get(report_type, "BÁO CÁO")
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center")
        
        # Thêm ngày xuất báo cáo
        ws.merge_cells('A2:G2')
        date_cell = ws['A2']
        date_cell.value = f"Ngày xuất báo cáo: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        date_cell.alignment = Alignment(horizontal="right")
        date_cell.font = Font(italic=True)
        
        # Thêm header
        for col_num, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Thêm dữ liệu
        for row_num, item in enumerate(data, 4):
            for col_num, column in enumerate(columns, 1):
                cell = ws.cell(row=row_num, column=col_num, value=item.get(column, ""))
                cell.border = thin_border
                if col_num in [4, 5]:  # Căn giữa cho cột số và ngày
                    cell.alignment = Alignment(horizontal="center")
        
        # Lưu file
        if not os.path.exists('reports'):
            os.makedirs('reports')
        
        filepath = os.path.join('reports', filename)
        wb.save(filepath)
        
        return filepath